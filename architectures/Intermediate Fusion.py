"""
Fusion Model for Multi-Modal Binary Classification of Manufacturing Defects

This script implements a deep learning fusion architecture that combines temporal image sequences
with ECT (Eddy Current Testing) sensor data for binary classification of manufacturing defects.
The model integrates a ResNet50 Temporal backbone for processing image sequences [t-2, t-1, t] with
an LSTM-based network enhanced by multi-head attention for ECT signal processing. The architecture
employs cross-modal attention mechanisms to enable interaction between visual and sensor modalities,
learnable weights for balancing modal contributions, and comprehensive feature selection. The
implementation includes adaptive device detection, stratified data splitting, class-weighted loss
functions, mixed precision training, and detailed per-detail performance analysis.
"""

import os
import torch
import torch.nn as nn
import torch.optim as optim
from torchvision import datasets, transforms
from torch.utils.data import DataLoader, Subset, Dataset
from sklearn.metrics import confusion_matrix, classification_report, precision_recall_fscore_support, roc_auc_score
import matplotlib.pyplot as plt
import seaborn as sns
import random
from tqdm import tqdm
import torchvision.models as models
import numpy as np
import warnings
import shutil
from PIL import Image
import pandas as pd
from sklearn.model_selection import train_test_split
from sklearn.feature_selection import mutual_info_classif
import multiprocessing
warnings.filterwarnings('ignore')

# ============================================================================
# PATH CONFIGURATION - MODIFY HERE TO CHANGE PATHS
# ============================================================================
# Path to folder containing all images (both compliant and defective)
IMAGES_DIR = r"C:\Users\NUM\Desktop\Coldani_Domini\fusion\dataset\data\total_images"

# Path to Excel file with temporal information (buildjob, geometry_name, channel, layer_index)
TEMPORAL_INFO_PATH = r"C:\Users\NUM\Desktop\fusion_final\images_sequences_labels_v1_POROSITY_final.xlsx"

# Path to Excel file with labels (image_name, label, detail)
LABELS_PATH = r"C:\Users\NUM\Desktop\fusion_final\final_labeling.xlsx"

# Path to Excel file with ECT data
ECT_PATH = r"C:\Users\NUM\Desktop\fusion_final\ECT_signal_final.xlsx"

# Base path for results
RESULTS_BASE_DIR = r"C:\Users\NUM\Desktop\trials_v1"

# Results folder (will be created automatically)
RESULTS_DIR = os.path.join(RESULTS_BASE_DIR, "251119-t3-fusion_model_v7_temporal_3steps_light_balanced")
# ============================================================================

# Protection for multiprocessing on Windows
multiprocessing.freeze_support()

# Classes must be defined at module level for multiprocessing on Windows
# -------------------------------
# IMPROVED FUSION DATASET - V7: IMAGE SEQUENCES + ECT
# -------------------------------
class FusionDataset(Dataset):
    """Dataset that combines image sequences and ECT data with detail support"""
    def __init__(self, image_dir, image_sequences_metadata, ect_sequences, transform=None):
        self.image_dir = image_dir
        self.transform = transform
        self.image_sequences_metadata = image_sequences_metadata
        self.ect_sequences = ect_sequences
        
        # OPTIMIZATION: Preload file list and cache image paths
        print("Preloading image directory...")
        try:
            self.all_files = set(f.lower() for f in os.listdir(self.image_dir) 
                               if f.lower().endswith(('.png', '.jpg', '.jpeg')))
            print(f"Preloaded {len(self.all_files)} image files")
        except Exception as e:
            print(f"Warning: Could not preload image directory: {e}")
            self.all_files = set()
        
        # Cache for image paths (image_name -> img_path)
        self.image_path_cache = {}
        
        # OPTIMIZATION: Cache loaded images in memory (for most used images)
        self.image_cache = {}  # image_name -> already transformed tensor
        self.cache_max_size = 1000  # Maximum number of images in cache
        
        print(f"FusionDataset initialized: {len(image_sequences_metadata)} sequences")
        defective_count = sum(1 for m in image_sequences_metadata if m['label'] == 0)
        compliant_count = sum(1 for m in image_sequences_metadata if m['label'] == 1)
        print(f"Distribution: Defective={defective_count}, Compliant={compliant_count}")
    
    def __len__(self):
        return len(self.image_sequences_metadata)
    
    def _load_image(self, image_name, use_cache=True):
        """Loads a single image with fallback - OPTIMIZED with cache"""
        # OPTIMIZATION: Check image cache before loading
        if use_cache and image_name in self.image_cache:
            return self.image_cache[image_name]
        
        # Use path cache if available
        if image_name in self.image_path_cache:
            img_path = self.image_path_cache[image_name]
        else:
            img_name_base = str(image_name).split('.')[0]
            
            # Search image with various methods
            search_paths = [
                os.path.join(self.image_dir, image_name),
                os.path.join(self.image_dir, image_name + '.png'),
                os.path.join(self.image_dir, image_name + '.PNG'),
                os.path.join(self.image_dir, img_name_base + '.png'),
                os.path.join(self.image_dir, img_name_base + '.PNG'),
            ]
            
            img_path = None
            for search_path in search_paths:
                if os.path.exists(search_path):
                    img_path = search_path
                    break
            
            # If not found, search by prefix using preloaded list
            if img_path is None:
                img_name_base_lower = img_name_base.lower()
                # Search in preloaded list (faster than os.listdir)
                possible_files = [f for f in self.all_files 
                                 if f.startswith(img_name_base_lower)]
                if possible_files:
                    # Take first match (restore original case if possible)
                    for orig_file in os.listdir(self.image_dir):
                        if orig_file.lower() == possible_files[0]:
                            img_path = os.path.join(self.image_dir, orig_file)
                            break
                    if img_path is None:
                        img_path = os.path.join(self.image_dir, possible_files[0])
            
            # If still not found, use as fallback
            if img_path is None:
                img_path = os.path.join(self.image_dir, image_name)
            
            # Save in cache for next calls
            self.image_path_cache[image_name] = img_path
        
        # Load image
        try:
            if not os.path.exists(img_path):
                image = Image.new('RGB', (512, 512), color=(0, 0, 0))
            else:
                image = Image.open(img_path).convert('RGB')
                if image.size[0] == 0 or image.size[1] == 0:
                    image = Image.new('RGB', (512, 512), color=(0, 0, 0))
        except Exception as e:
            image = Image.new('RGB', (512, 512), color=(0, 0, 0))
        
        # Apply transformations
        if self.transform:
            image = self.transform(image)
        else:
            # If transform is None, apply at least resize and conversion to tensor
            # This is necessary when dataset is used without transform (e.g. during split)
            transform_default = transforms.Compose([
                transforms.Resize((384, 384)),  # Use img_size by default, but 384 is a safe value
                transforms.ToTensor(),
                transforms.Normalize([0.485, 0.456, 0.406], [0.229, 0.224, 0.225])
            ])
            image = transform_default(image)
        
        # OPTIMIZATION: Save in cache if not in training (transform changes every time)
        if use_cache and self.transform is None and len(self.image_cache) < self.cache_max_size:
            self.image_cache[image_name] = image
        
        return image
    
    def __getitem__(self, idx):
        meta = self.image_sequences_metadata[idx]
        sequence_images = meta['sequence_images']  # [t-2, t-1, t]
        ect_idx = meta['ect_sequence_idx']
        label = meta['label']
        detail = meta['detail']
        center_image = meta['center_image']
        
        # Load image sequence [t-2, t-1, t]
        image_sequence = []
        for img_name in sequence_images:
            image = self._load_image(img_name)
            image_sequence.append(image)
        
        # Stack images: (sequence_length, C, H, W)
        image_sequence_tensor = torch.stack(image_sequence)
        
        # Load ECT sequence
        ect_sequence = torch.FloatTensor(self.ect_sequences[ect_idx])
        
        # Check and clean NaN/Inf in ECT data
        if torch.isnan(ect_sequence).any() or torch.isinf(ect_sequence).any():
            # Replace NaN/Inf with 0
            ect_sequence = torch.nan_to_num(ect_sequence, nan=0.0, posinf=0.0, neginf=0.0)
        
        return image_sequence_tensor, ect_sequence, label, center_image, detail

# -------------------------------
# IMPROVED FUSION MODEL WITH TEMPORAL RESNET
# -------------------------------
class TemporalResNetModel(nn.Module):
    """ResNet50 with temporal aggregation for image sequences (from ResNet50_temporal_v2)"""
    def __init__(self, num_classes=2, pretrained=True, model_name='resnet50', 
                 sequence_length=3, hidden_size=256, num_lstm_layers=2, dropout=0.3):
        super(TemporalResNetModel, self).__init__()
        
        self.sequence_length = sequence_length
        
        # ResNet50 backbone (without final classifier)
        if model_name == 'resnet50':
            backbone = models.resnet50(pretrained=pretrained)
            num_features = backbone.fc.in_features
            self.backbone = nn.Sequential(*list(backbone.children())[:-1])
        else:
            backbone = models.resnet50(pretrained=pretrained)
            num_features = backbone.fc.in_features
            self.backbone = nn.Sequential(*list(backbone.children())[:-1])
        
        # LSTM for temporal aggregation
        self.lstm = nn.LSTM(
            input_size=num_features,
            hidden_size=hidden_size,
            num_layers=num_lstm_layers,
            batch_first=True,
            dropout=dropout if num_lstm_layers > 1 else 0,
            bidirectional=True
        )
        
        # Multi-Head Attention
        self.attention = nn.MultiheadAttention(
            embed_dim=hidden_size * 2,  # *2 for bidirectional
            num_heads=8,
            dropout=dropout,
            batch_first=True
        )
        
        # Layer Normalization
        self.layer_norm = nn.LayerNorm(hidden_size * 2)
        
        # Attention pooling for temporal aggregation (V2 - BETTER PERFORMANCE)
        self.attention_pool = nn.MultiheadAttention(
            embed_dim=hidden_size * 2,
            num_heads=4,
            dropout=dropout,
            batch_first=True
        )
        
        # Query for attention pooling (uses central timestep as query)
        self.pool_query = nn.Parameter(torch.randn(1, 1, hidden_size * 2))
        
        # Output size for fusion
        self.output_size = hidden_size * 2
    
    def forward(self, x):
        # x shape: (batch_size, sequence_length, C, H, W)
        batch_size, seq_len, C, H, W = x.size()
        
        # OPTIMIZATION: More efficient reshape using contiguous
        x_reshaped = x.view(batch_size * seq_len, C, H, W).contiguous()
        
        # Extract features with ResNet for each image
        features = self.backbone(x_reshaped)  # (batch_size * seq_len, num_features, 1, 1)
        # OPTIMIZATION: More efficient flatten
        features = features.flatten(1)  # (batch_size * seq_len, num_features)
        
        # OPTIMIZATION: More efficient reshape
        features_seq = features.view(batch_size, seq_len, -1).contiguous()
        
        # LSTM for temporal aggregation
        lstm_out, _ = self.lstm(features_seq)  # (batch_size, seq_len, hidden_size*2)
        
        # Multi-Head Attention
        attn_out, attn_weights = self.attention(lstm_out, lstm_out, lstm_out)
        
        # Residual connection + Layer Normalization
        residual_out = self.layer_norm(lstm_out + attn_out)
        
        # V2: Attention Pooling instead of simple mean (BETTER PERFORMANCE)
        # Uses central timestep as query to weight all sequences
        query = self.pool_query.expand(batch_size, -1, -1)  # (batch_size, 1, hidden_size*2)
        pooled_out, pool_weights = self.attention_pool(query, residual_out, residual_out)
        aggregated = pooled_out.squeeze(1)  # (batch_size, hidden_size*2)
        
        return aggregated

class FusionModel(nn.Module):
    """BALANCED fusion model: TemporalResNet50 (image sequences) + LSTM+Attention (ECT) + Cross-Modal Attention + Weight Balancing"""
    def __init__(self, num_classes=2, dropout_rate=0.4, ect_input_size=20, sequence_length=3):
        super(FusionModel, self).__init__()
        
        # TemporalResNet for image sequences (from ResNet50_temporal_v2)
        self.temporal_resnet = TemporalResNetModel(
            num_classes=num_classes,
            pretrained=True,
            model_name='resnet50',
            sequence_length=sequence_length,
            hidden_size=256,
            num_lstm_layers=2,
            dropout=0.3
        )
        
        self.image_features_size = self.temporal_resnet.output_size  # 512 (256*2)
        
        # LSTM+Attention for ECT data
        hidden_size = 128
        num_layers = 2
        num_heads = 8
        
        self.lstm = nn.LSTM(ect_input_size, hidden_size, num_layers, 
                           batch_first=True, dropout=0.2, bidirectional=True)
        
        self.attention = nn.MultiheadAttention(
            embed_dim=hidden_size * 2,
            num_heads=num_heads,
            dropout=0.2,
            batch_first=True
        )
        
        self.layer_norm = nn.LayerNorm(hidden_size * 2)
        self.ect_features_size = hidden_size * 2  # 256
        
        # Normalization to balance TemporalResNet and LSTM
        self.image_norm = nn.LayerNorm(self.image_features_size)
        self.ect_norm = nn.LayerNorm(self.ect_features_size)
        
        # Projection to balance dimensions
        self.image_projection = nn.Sequential(
            nn.Linear(self.image_features_size, 512),
            nn.BatchNorm1d(512),
            nn.ReLU(inplace=True),
            nn.Dropout(0.3)
        )
        
        # Projection for LSTM too to have similar dimensions
        self.ect_projection = nn.Sequential(
            nn.Linear(self.ect_features_size, 512),  # Brings to 512 like TemporalResNet
            nn.BatchNorm1d(512),
            nn.ReLU(inplace=True),
            nn.Dropout(0.2)
        )
        
        # ============================================================================
        # CROSS-MODAL ATTENTION: Allows the two modalities to interact
        # ============================================================================
        # Common dimension after projection
        common_dim = 512
        
        # Cross-modal attention: images "attend to" ECT
        self.cross_attention_img_to_ect = nn.MultiheadAttention(
            embed_dim=common_dim,
            num_heads=8,
            dropout=0.2,
            batch_first=True
        )
        
        # Cross-modal attention: ECT "attend to" images
        self.cross_attention_ect_to_img = nn.MultiheadAttention(
            embed_dim=common_dim,
            num_heads=8,
            dropout=0.2,
            batch_first=True
        )
        
        # Layer normalization for cross-modal features
        self.cross_norm_img = nn.LayerNorm(common_dim)
        self.cross_norm_ect = nn.LayerNorm(common_dim)
        
        # Final projection after cross-attention
        self.image_cross_projection = nn.Sequential(
            nn.Linear(common_dim, 512),
            nn.BatchNorm1d(512),
            nn.ReLU(inplace=True),
            nn.Dropout(0.2)
        )
        
        self.ect_cross_projection = nn.Sequential(
            nn.Linear(common_dim, 512),
            nn.BatchNorm1d(512),
            nn.ReLU(inplace=True),
            nn.Dropout(0.2)
        )
        
        # ============================================================================
        # LEARNABLE WEIGHT BALANCING: Weights to balance modal contributions
        # ============================================================================
        # Learnable weights to balance contributions (initialized to 0.5 for both)
        # We use sigmoid to keep weights between 0 and 1, and normalize for sum = 1
        self.modal_weights = nn.Parameter(torch.tensor([0.5, 0.5]))  # [img_weight, ect_weight]
        
        # Learnable gate to control how much to use original vs cross-modal features
        self.cross_gate_img = nn.Parameter(torch.tensor(0.5))  # 0 = only original, 1 = only cross
        self.cross_gate_ect = nn.Parameter(torch.tensor(0.5))
        
        # Common classifier (after concatenation)
        # Now both branches have 512 features, so total 1024
        fused_features_size = 512 + 512  # 1024 (TemporalResNet 512 + LSTM 512)
        
        self.classifier = nn.Sequential(
            nn.Linear(fused_features_size, 512),
            nn.BatchNorm1d(512),
            nn.ReLU(inplace=True),
            nn.Dropout(dropout_rate),
            nn.Linear(512, 256),
            nn.BatchNorm1d(256),
            nn.ReLU(inplace=True),
            nn.Dropout(dropout_rate * 0.8),
            nn.Linear(256, num_classes)
        )
    
    def forward(self, image_sequences, ect_sequences):
        # Feature extraction for image sequences (TemporalResNet)
        # image_sequences: (batch, sequence_length, C, H, W)
        img_features = self.temporal_resnet(image_sequences)  # (batch, 512)
        img_features = self.image_norm(img_features)
        img_features = self.image_projection(img_features)  # (batch, 512)
        
        # Feature extraction for ECT (LSTM + Attention)
        lstm_out, _ = self.lstm(ect_sequences)  # (batch, seq_len, 256)
        attn_out, _ = self.attention(lstm_out, lstm_out, lstm_out)
        residual_out = self.layer_norm(lstm_out + attn_out)
        ect_features = residual_out.mean(dim=1)  # (batch, 256)
        ect_features = self.ect_norm(ect_features)
        ect_features = self.ect_projection(ect_features)  # (batch, 512)
        
        # ============================================================================
        # CROSS-MODAL ATTENTION: Interaction between the two modalities
        # ============================================================================
        # Reshape for attention: (batch, 1, 512) to enable attention
        img_features_expanded = img_features.unsqueeze(1)  # (batch, 1, 512)
        ect_features_expanded = ect_features.unsqueeze(1)  # (batch, 1, 512)
        
        # Cross-attention: images "attend to" ECT
        # Query = img_features, Key/Value = ect_features
        img_cross_attn, _ = self.cross_attention_img_to_ect(
            img_features_expanded,  # query
            ect_features_expanded,   # key
            ect_features_expanded    # value
        )
        img_cross_attn = img_cross_attn.squeeze(1)  # (batch, 512)
        img_cross_attn = self.cross_norm_img(img_cross_attn)
        
        # Cross-attention: ECT "attend to" images
        # Query = ect_features, Key/Value = img_features
        ect_cross_attn, _ = self.cross_attention_ect_to_img(
            ect_features_expanded,   # query
            img_features_expanded,   # key
            img_features_expanded    # value
        )
        ect_cross_attn = ect_cross_attn.squeeze(1)  # (batch, 512)
        ect_cross_attn = self.cross_norm_ect(ect_cross_attn)
        
        # Gate to balance original vs cross-modal features
        # Sigmoid to keep gate between 0 and 1
        gate_img = torch.sigmoid(self.cross_gate_img)
        gate_ect = torch.sigmoid(self.cross_gate_ect)
        
        # Combination of original + cross-modal features
        img_features_enhanced = gate_img * img_cross_attn + (1 - gate_img) * img_features
        ect_features_enhanced = gate_ect * ect_cross_attn + (1 - gate_ect) * ect_features
        
        # Final projection after cross-attention
        img_features_final = self.image_cross_projection(img_features_enhanced)  # (batch, 512)
        ect_features_final = self.ect_cross_projection(ect_features_enhanced)  # (batch, 512)
        
        # ============================================================================
        # LEARNABLE WEIGHT BALANCING: Weights to balance contributions
        # ============================================================================
        # Normalize weights for sum = 1 using softmax
        modal_weights_normalized = torch.softmax(self.modal_weights, dim=0)
        w_img = modal_weights_normalized[0]
        w_ect = modal_weights_normalized[1]
        
        # Apply weights to features before concatenation
        # This forces the model to balance contributions
        img_features_weighted = img_features_final * w_img
        ect_features_weighted = ect_features_final * w_ect
        
        # Balanced concatenation with learnable weights
        fused_features = torch.cat([img_features_weighted, ect_features_weighted], dim=1)  # (batch, 1024)
        output = self.classifier(fused_features)
        
        return output

# -------------------------------
# EARLY STOPPING
# -------------------------------
class EarlyStopping:
    def __init__(self, patience=5, min_delta=0.005, restore_best_weights=True):
        self.patience = patience
        self.min_delta = min_delta
        self.restore_best_weights = restore_best_weights
        self.counter = 0
        self.best_loss = None
        self.best_acc = None
        self.best_model_state = None
        self.epochs_without_improvement = 0

    def __call__(self, val_loss, val_acc, model):
        if self.best_loss is None:
            self.best_loss = val_loss
            self.best_acc = val_acc
            self.best_model_state = model.state_dict().copy()
            self.epochs_without_improvement = 0
        elif val_loss < self.best_loss - self.min_delta:
            self.best_loss = val_loss
            self.best_acc = val_acc
            self.counter = 0
            self.epochs_without_improvement = 0
            self.best_model_state = model.state_dict().copy()
        else:
            self.counter += 1
            self.epochs_without_improvement += 1
        
        if self.epochs_without_improvement >= self.patience:
            if self.restore_best_weights:
                model.load_state_dict(self.best_model_state)
            return True
        return False

# -------------------------------
# CUSTOM COLLATE FUNCTION
# -------------------------------
def custom_collate_fn(batch):
    """Collate function for image sequences - must be at module level for multiprocessing"""
    image_sequences = []
    ect_seqs = []
    labels = []
    center_images = []
    details = []
    
    for img_seq, ect_seq, label, center_img, detail in batch:
        # img_seq is already (sequence_length, C, H, W) - stacked in dataset
        image_sequences.append(img_seq)
        ect_seqs.append(ect_seq)
        labels.append(label)
        center_images.append(center_img)
        details.append(detail)
    
    # Stack image sequences: (batch, sequence_length, C, H, W)
    # Images are already resized to fixed size, so no padding needed
    image_sequences = torch.stack(image_sequences)
    ect_seqs = torch.stack(ect_seqs)
    labels = torch.tensor(labels, dtype=torch.long)
    
    return image_sequences, ect_seqs, labels, center_images, details

# ============================================================================
# MAIN CODE
# ============================================================================
if __name__ == '__main__':

    # -------------------------------
    # 1. INTELLIGENT DEVICE SETUP
    # -------------------------------
    def setup_device():
        """Intelligent device setup with automatic GPU/CPU detection"""
        print("="*80)
        print("HARDWARE DETECTION")
        print("="*80)
        
        if torch.cuda.is_available():
            device = torch.device("cuda")
            gpu_name = torch.cuda.get_device_name(0)
            gpu_memory = torch.cuda.get_device_properties(0).total_memory / 1024**3
            
            print(f"GPU DETECTED: {gpu_name}")
            print(f"GPU Memory: {gpu_memory:.1f} GB")
            print(f"CUDA Version: {torch.version.cuda}")
            
            try:
                test_tensor = torch.randn(100, 100).to(device)
                result = torch.mm(test_tensor, test_tensor)
                print("GPU Test: SUCCESS")
                return device, "gpu", gpu_memory
            except Exception as e:
                raise RuntimeError(f"  GPU Test: FAILED - {str(e)}")
        else:
            raise RuntimeError("❌ GPU not available! Model requires GPU.")

    # -------------------------------
    # 2. ADAPTIVE PARAMETERS
    # -------------------------------
    def get_adaptive_params(device_type, gpu_memory=0):
        """Optimized parameters based on available device"""
        
        if device_type == "gpu":
            if gpu_memory >= 8:
                batch_size = 32
                num_workers = 0  # Set to 0 to avoid multiprocessing issues (like v0/v6)
                img_size = 384
            elif gpu_memory >= 4:
                batch_size = 24
                num_workers = 0  # Set to 0 to avoid multiprocessing issues (like v0/v6)
                img_size = 384
            else:
                batch_size = 16
                num_workers = 0  # Set to 0 to avoid multiprocessing issues (like v0/v6)
                img_size = 384
            
            pin_memory = True
            num_epochs = 30  # INCREASED from 20 to 30
            
        else:  # CPU
            batch_size = 16
            num_workers = 0  # Set to 0 to avoid multiprocessing issues (like v0/v6)
            img_size = 384
            pin_memory = False
            num_epochs = 30  # INCREASED from 15 to 30
            print("CPU MODE")
        
        # Common parameters - REDUCED REGULARIZATION (V6)
        learning_rate = 1e-4  # INCREASED from 5e-5 to 1e-4
        weight_decay = 1e-4  # REDUCED from 1e-3 to 1e-4
        dropout_rate = 0.4  # REDUCED from 0.6 to 0.4
        
        print(f"Parameters: Batch={batch_size}, Workers={num_workers}, Size={img_size}")
        print(f"Epochs: {num_epochs}, Pin Memory: {pin_memory}")
        print(f"REDUCED Regularization (V6): LR={learning_rate}, WD={weight_decay}, Dropout={dropout_rate}")
          
        return {
            'batch_size': batch_size,
            'num_workers': num_workers,
            'img_size': img_size,
            'pin_memory': pin_memory,
            'num_epochs': num_epochs,
            'learning_rate': learning_rate,
            'weight_decay': weight_decay,
            'dropout_rate': dropout_rate
        }

    # -------------------------------
    # 3. ECT DATA PREPROCESSING
    # -------------------------------
    def extract_features_per_geometry(df):
        """Extract features for each geometry, channel and buildjob - MODIFIED LOGIC"""
        print("Extracting features per geometry, channel and buildjob...")
        
        # Keep only required columns and ignore everything else
        required_cols = ['real', 'imag', 'channel', 'layer_index', 'geometry_name', 'buildjob', 'corresponding_image']
        present_cols = [c for c in required_cols if c in df.columns]
        missing_cols = [c for c in required_cols if c not in df.columns]
        if missing_cols:
            print(f"WARNING: Missing columns in ECT dataframe: {missing_cols}")
        df = df[present_cols].copy()
        
        all_features = []
        
        # Build grouping keys using available columns among required ones
        groupby_cols = [c for c in ['geometry_name', 'channel', 'buildjob', 'layer_index', 'corresponding_image'] if c in df.columns]
        if 'buildjob' not in groupby_cols:
            print("WARNING: Column 'buildjob' not found! Grouping will proceed without 'buildjob'")
        if 'corresponding_image' not in groupby_cols:
            print("WARNING: Column 'corresponding_image' not found! Grouping will proceed without 'corresponding_image'")
        
        for group_key, group in df.groupby(groupby_cols):
            # Extract known fields with safe fallback
            key_map = dict(zip(groupby_cols, group_key if isinstance(group_key, tuple) else (group_key,)))
            geometry = key_map.get('geometry_name', None)
            channel = key_map.get('channel', None)
            buildjob = key_map.get('buildjob', None)
            layer = key_map.get('layer_index', None)
            corr_image = key_map.get('corresponding_image', group['corresponding_image'].iloc[0] if 'corresponding_image' in group.columns else None)
            
            layer_data = group
            real_values = layer_data['real'].values
            imag_values = layer_data['imag'].values
            
            if len(real_values) == 0 or len(imag_values) == 0:
                continue
            
            # Label is no longer read from ECT file, it is taken from final_labels.xlsx
            # We only keep placeholders for layer_type and Defect, they will be replaced after matching
            features = {
                'geometry_name': geometry,
                'channel': channel,
                'buildjob': buildjob if buildjob is not None else (layer_data['buildjob'].iloc[0] if 'buildjob' in layer_data.columns else None),
                'layer_index': layer,
                'corresponding_image': corr_image,
                'layer_type': 'unknown',  # Verrà determinato dal matching con labels
                'Defect': 0,  # Placeholder, verrà determinato dal matching con labels
                
                # Features real
                'mean_real': np.mean(real_values),
                'std_real': np.std(real_values),
                'var_real': np.var(real_values),
                'skewness_real': pd.Series(real_values).skew() if len(real_values) > 2 and np.std(real_values) > 0 else 0.0,
                'kurtosis_real': pd.Series(real_values).kurtosis() if len(real_values) > 2 and np.std(real_values) > 0 else 0.0,
                
                # Features imag
                'mean_imag': np.mean(imag_values),
                'std_imag': np.std(imag_values),
                'var_imag': np.var(imag_values),
                'skewness_imag': pd.Series(imag_values).skew() if len(imag_values) > 2 and np.std(imag_values) > 0 else 0.0,
                'kurtosis_imag': pd.Series(imag_values).kurtosis() if len(imag_values) > 2 and np.std(imag_values) > 0 else 0.0,
                
                # Features FFT
                'fft_real': np.abs(np.fft.fft(real_values)).mean(),
                'fft_imag': np.abs(np.fft.fft(imag_values)).mean(),
                
                # Features gradient
                'gradient_real': np.gradient(real_values).mean() if len(real_values) > 1 else 0.0,
                'gradient_imag': np.gradient(imag_values).mean() if len(imag_values) > 1 else 0.0,
                
                # Features peaks and valleys
                'peaks_real': len(np.where(np.diff(np.sign(np.diff(real_values))) < 0)[0]) if len(real_values) > 2 else 0.0,
                'valleys_real': len(np.where(np.diff(np.sign(np.diff(real_values))) > 0)[0]) if len(real_values) > 2 else 0.0,
                'peaks_imag': len(np.where(np.diff(np.sign(np.diff(imag_values))) < 0)[0]) if len(imag_values) > 2 else 0.0,
                'valleys_imag': len(np.where(np.diff(np.sign(np.diff(imag_values))) > 0)[0]) if len(imag_values) > 2 else 0.0,
                
                # Features KernelPCA (placeholder)
                'kpca_real': 0.0,
                'kpca_imag': 0.0,
                
                # Features energy and power
                'energy_real': np.sum(real_values**2),
                'energy_imag': np.sum(imag_values**2),
                'power_real': np.mean(real_values**2),
                'power_imag': np.mean(imag_values**2),
            }
            
            all_features.append(features)
        
        features_df = pd.DataFrame(all_features)
        print(f"Features extracted: {len(features_df)} layer sequences")
        
        return features_df

    def create_centered_temporal_sequences(features_df, sequence_length=3):
        """Create CENTERED temporal sequences (2 before + center) - ORIGINAL LOGIC"""
        print(f"Creating CENTERED temporal sequences (length: {sequence_length})...")
        print("Pattern: [t-2, t-1, t] where t is the layer to classify")
        
        sequences = []
        metadata = []
        
        # Group temporal sequences: to build windows on layer_index
        # we use wide groups (without layer_index and without corresponding_image) and sort by layer_index
        if 'buildjob' in features_df.columns:
            groupby_cols = ['geometry_name', 'channel', 'buildjob']
        else:
            groupby_cols = ['geometry_name', 'channel']
            print("WARNING: Column 'buildjob' not found! Grouping only by geometry and channel")
        print(f"Grouping for sequences: {', '.join(groupby_cols)}")
        
        # Count total groups for progress
        grouped = list(features_df.groupby(groupby_cols))
        total_groups = len(grouped)
        print(f"Total groups to process: {total_groups}")
        
        for group_idx, (group_key, group) in enumerate(grouped):
            if (group_idx + 1) % 100 == 0 or group_idx == 0:
                print(f"Processing group {group_idx + 1}/{total_groups}...")
            # Dynamic unpack of group key
            geometry = group_key[0]
            channel = group_key[1]
            buildjob = None
            if 'buildjob' in groupby_cols:
                buildjob = group_key[2]
            # Note: if present, 'corresponding_layer' is the last element of the key
            group_sorted = group.sort_values('layer_index')
            layers = group_sorted['layer_index'].values
            
            # Find consecutive sequences
            consecutive_groups = []
            current_group = [layers[0]]
            
            for i in range(1, len(layers)):
                if layers[i] == layers[i-1] + 1:  # Consecutive
                    current_group.append(layers[i])
                else:  # Gap found
                    if len(current_group) >= sequence_length:
                        consecutive_groups.append(current_group)
                    current_group = [layers[i]]
            
            # Add the last group if it's long enough
            if len(current_group) >= sequence_length:
                consecutive_groups.append(current_group)
            
            # Create CENTERED sequences from each consecutive group
            for seq_group_idx, group_layers in enumerate(consecutive_groups):
                if (group_idx + 1) % 100 == 0 and seq_group_idx == 0:
                    print(f"   Creating sequences for consecutive group with {len(group_layers)} layers...")
                group_data = group_sorted[group_sorted['layer_index'].isin(group_layers)]
                feature_cols = [col for col in group_data.columns if col not in 
                             ['geometry_name', 'channel', 'buildjob', 'layer_index', 'corresponding_layer', 'corresponding_image', 'layer_type', 'Defect']]
                
                # Create sequences for ALL layers, including edges
                n_layers_in_group = len(group_layers)
                if n_layers_in_group > 100 and seq_group_idx == 0:
                    print(f"   Creating {n_layers_in_group} sequences (this may take a moment)...")
                for i in range(n_layers_in_group):  # Now includes edges too!
                    # CENTERED pattern: [t-2, t-1, t]
                    # For edges, use padding (repeat first/last available layer)
                    sequence_layers_idx = []
                    for offset in [-2, -1, 0]:
                        idx = i + offset
                        if idx < 0:
                            idx = 0  # Padding: use first layer
                        elif idx >= len(group_layers):
                            idx = len(group_layers) - 1  # Padding: use last layer
                        sequence_layers_idx.append(idx)
                    
                    sequence_layers = [group_layers[idx] for idx in sequence_layers_idx]
                    
                    sequence_data = group_data[group_data['layer_index'].isin(sequence_layers)]
                    sequence_data = sequence_data.sort_values('layer_index')
                    
                    # Extract features maintaining order [t-2, t-1, t]
                    sequence_features_list = []
                    for layer_idx in sequence_layers:
                        layer_row = sequence_data[sequence_data['layer_index'] == layer_idx]
                        if len(layer_row) > 0:
                            sequence_features_list.append(layer_row[feature_cols].values[0])
                        else:
                            # Fallback (should not happen)
                            sequence_features_list.append(group_data.iloc[0][feature_cols].values)
                    
                    sequence_features = np.array(sequence_features_list)
                    try:
                        corr_image = group_data.iloc[i]['corresponding_image']  # Current layer (image)
                    except (IndexError, KeyError):
                        # Fallback if cannot find corresponding_image
                        if len(group_data) > 0:
                            corr_image = group_data.iloc[min(i, len(group_data)-1)].get('corresponding_image', None)
                        else:
                            corr_image = None
                    
                    sequences.append(sequence_features)
                    metadata.append({
                        'corresponding_image': corr_image
                    })
        
        sequences = np.array(sequences)
        
        print(f"\nCENTERED sequences created: {len(sequences)}")
        print(f"Sequence shape: {sequences.shape}")
        
        return sequences, metadata

    # -------------------------------
    # 4. CREATION OF IMAGE SEQUENCES ALIGNED WITH ECT
    # -------------------------------
    def create_image_sequences_aligned_with_ect(ect_sequences, ect_metadata, temporal_dict, labels_dict, details_dict, image_dir):
        """Creates image sequences aligned with ECT sequences (same layer_index)"""
        print(f"Creating image sequences aligned with ECT sequences...")
        print("Pattern: image sequences must have the same layer_index as ECT sequences")
        
        # Helper function to normalize names
        def normalize_name(name):
            if pd.isna(name):
                return None
            name_str = str(name).strip()
            for ext in ['.png', '.PNG', '.jpg', '.JPG', '.jpeg', '.JPEG']:
                if name_str.lower().endswith(ext.lower()):
                    name_str = name_str[:-len(ext)]
                    break
            return name_str
        
        # Create mapping image_name (normalized) -> (original name, temporal information)
        image_temporal_map = {}
        for image_name, temp_info in temporal_dict.items():
            normalized = normalize_name(image_name)
            if normalized is not None:
                image_temporal_map[normalized] = (image_name, temp_info)  # Save original name and info
        
        # Group images by (buildjob, geometry_name, channel) and reorder by layer_index
        # As in ResNet50_temporal_v2
        print("Grouping and reordering images by (buildjob, geometry, channel)...")
        grouped_images = {}
        for norm_name, (orig_name, temp_info) in image_temporal_map.items():
            buildjob = temp_info.get('buildjob', None)
            geometry = temp_info.get('geometry_name', None)
            channel = temp_info.get('channel', None)
            layer_idx = temp_info.get('layer_index', None)
            
            if buildjob is None or geometry is None or channel is None or layer_idx is None:
                continue
            
            key = (buildjob, geometry, channel)
            if key not in grouped_images:
                grouped_images[key] = []
            
            grouped_images[key].append({
                'image_name': orig_name,
                'image_name_norm': norm_name,
                'layer_index': layer_idx
            })
        
        # Reorder by layer_index within each group
        for key in grouped_images:
            grouped_images[key] = sorted(grouped_images[key], key=lambda x: x['layer_index'])
        
        print(f"Grouped {len(grouped_images)} image groups")
        
        # Create mapping for fast search: (buildjob, geometry, channel, layer) -> original name
        layer_to_image = {}
        for norm_name, (orig_name, temp_info) in image_temporal_map.items():
            key = (temp_info.get('buildjob'), temp_info.get('geometry_name'), 
                   temp_info.get('channel'), temp_info.get('layer_index'))
            if key not in layer_to_image:
                layer_to_image[key] = orig_name  # Use original name
        
        # Create image sequences aligned with ECT
        # For each ECT sequence, use the same layer_index
        image_sequences_metadata = []
        valid_indices = []
        
        label_to_num = {'defective': 0, 'compliant': 1}
        
        print("Creating image sequences aligned with ECT sequences...")
        for i, ect_meta in enumerate(ect_metadata):
            center_image = ect_meta.get('corresponding_image', None)
            if center_image is None:
                continue
            
            center_image_norm = normalize_name(center_image)
            if center_image_norm not in image_temporal_map:
                continue
            
            # Get temporal information of central image
            center_original_name, center_temp = image_temporal_map[center_image_norm]
            buildjob = center_temp.get('buildjob', None)
            geometry = center_temp.get('geometry_name', None)
            channel = center_temp.get('channel', None)
            center_layer = center_temp.get('layer_index', None)
            
            if buildjob is None or geometry is None or channel is None or center_layer is None:
                continue
            
            # Extract layer_index from corresponding ECT sequence
            # ECT sequence has pattern [t-2, t-1, t] where t = center_layer
            # So layer_index are: [center_layer-2, center_layer-1, center_layer]
            ect_layer_indices = [
                center_layer - 2,  # t-2
                center_layer - 1,  # t-1
                center_layer       # t (center)
            ]
            
            # Search images corresponding to the same layer_index as ECT sequence
            sequence_images = []
            all_found = True
            
            for layer_idx in ect_layer_indices:
                key = (buildjob, geometry, channel, layer_idx)
                found_image = layer_to_image.get(key, None)
                
                if found_image is None:
                    # If not found, use central image as fallback (padding)
                    found_image = center_original_name
                    all_found = False
                
                sequence_images.append(found_image)
            
            # Verify that central image has label
            if center_image_norm not in labels_dict:
                continue
            
            label_str = labels_dict[center_image_norm]
            label = label_to_num.get(label_str.lower().strip(), None)
            if label is None:
                continue
            
            detail = details_dict.get(center_image_norm, 'unknown')
            if pd.isna(detail):
                detail = 'unknown'
            else:
                detail = str(detail).strip()
            
            image_sequences_metadata.append({
                'sequence_images': sequence_images,  # [t-2, t-1, t] - same layer_index as ECT
                'sequence_layers': ect_layer_indices,  # Sequence layer indices
                'center_image': center_image_norm,
                'label': label,
                'detail': detail,
                'buildjob': buildjob,
                'geometry': geometry,
                'channel': channel,
                'center_layer': center_layer,
                'ect_sequence_idx': i  # Index of corresponding ECT sequence
            })
            valid_indices.append(i)
        
        print(f"Created {len(image_sequences_metadata)} aligned image sequences")
        print(f"Valid sequences: {len(valid_indices)}/{len(ect_metadata)}")
        print(f"Pattern: image sequences use the same layer_index as ECT sequences")
        
        return image_sequences_metadata, valid_indices

    # Classes FusionDataset, TemporalResNetModel, FusionModel and EarlyStopping
    # are already defined at module level (before if __name__ == '__main__' block)
    # to allow multiprocessing on Windows

    # -------------------------------
    # 7. MIXUP REMOVED (V6)
    # -------------------------------
    # Mixup completely removed in V6 to improve accuracy

    # -------------------------------
    # 8. FEATURE SELECTION
    # -------------------------------
    def feature_selection(X, y, top_k=16):  # INCREASED from 14 to 16
        """Select top features using mutual information"""
        print(f"Selecting top {top_k} features using mutual information...")
        
        # Handle NaN values
        X = X.fillna(0)
        
        # Calculate mutual information
        mi_scores = mutual_info_classif(X, y)
        
        # Get top features
        top_features_idx = np.argsort(mi_scores)[-top_k:]
        top_features_names = X.columns[top_features_idx]
        
        print(f"Top {top_k} features:")
        for i, (idx, name) in enumerate(zip(top_features_idx, top_features_names)):
            print(f"   {i+1:2d}. {name}: {mi_scores[idx]:.4f}")
        
        return top_features_idx, top_features_names

    # -------------------------------
    # MAIN
    # -------------------------------
def main():
    print("\nSTARTING MAIN EXECUTION...")
    # Non-deterministic execution: no fixed seeds
    torch.backends.cudnn.deterministic = False
    torch.backends.cudnn.benchmark = True
    
    print("="*80)
    print("FUSION MODEL V7 BALANCED (3 STEPS): TemporalResNet50 + LSTM+Attention + Cross-Modal Attention + Learnable Weights")
    print("="*80)
    
    # Setup device and parameters
    device, device_type, gpu_memory = setup_device()
    params = get_adaptive_params(device_type, gpu_memory)
    batch_size = params['batch_size']
    num_epochs = params['num_epochs']
    learning_rate = params['learning_rate']
    img_size = params['img_size']
    weight_decay = params['weight_decay']
    dropout_rate = params['dropout_rate']
    num_workers = params['num_workers']
    pin_memory = params['pin_memory']
    
    # Use paths from configuration at top
    images_dir = IMAGES_DIR
    temporal_info_path = TEMPORAL_INFO_PATH
    labels_path = LABELS_PATH
    ect_path = ECT_PATH
    results_dir = RESULTS_DIR
    
    # Verify file and directory existence
    print("\nVerifying file paths...")
    if not os.path.exists(images_dir):
        raise FileNotFoundError(f"Image directory not found: {images_dir}")
    print(f"Image directory: {images_dir}")
    
    if not os.path.exists(temporal_info_path):
        raise FileNotFoundError(f"Temporal information file not found: {temporal_info_path}")
    print(f"Temporal information file: {temporal_info_path}")
    
    if not os.path.exists(labels_path):
        raise FileNotFoundError(f"Labels file not found: {labels_path}")
    print(f"Labels file: {labels_path}")
    
    if not os.path.exists(ect_path):
        raise FileNotFoundError(f"ECT file not found: {ect_path}")
    print(f"ECT file: {ect_path}")
    
    # Folder to save all outputs
    if not os.path.exists(results_dir):
        os.makedirs(results_dir)
        print(f"Created results directory: {results_dir}")
    else:
        print(f"Using existing results directory: {results_dir}")
    
    # Load temporal information (buildjob, geometry_name, channel, layer_index)
    print("\nLoading temporal information...")
    try:
        # Try different engines to avoid issues with Python 3.13
        temporal_df = None
        engines_to_try = ['calamine', 'openpyxl', 'xlrd', None]  # calamine is more stable with Python 3.13
        
        for engine in engines_to_try:
            try:
                if engine is None:
                    print(f"   Attempting with default engine...")
                    temporal_df = pd.read_excel(temporal_info_path)
                else:
                    print(f"   Attempting with engine '{engine}'...")
                    temporal_df = pd.read_excel(temporal_info_path, engine=engine)
                print(f"   Loading completed with engine '{engine if engine else 'default'}'")
                break
            except ImportError:
                print(f"   Engine '{engine}' not available, trying next...")
                continue
            except Exception as e:
                print(f"   Engine '{engine}' failed: {str(e)[:100]}...")
                if engine == engines_to_try[-1]:  # Last attempt
                    raise
                continue
        
        if temporal_df is None:
            raise RuntimeError("Unable to load Excel file with any available engine")
        
        print(f"Temporal info loaded: {len(temporal_df)} samples")
        print(f"Columns: {list(temporal_df.columns)}")
        
        # Verify required columns
        required_cols = ['image_name', 'buildjob', 'geometry_name', 'channel', 'layer_index']
        missing_cols = [col for col in required_cols if col not in temporal_df.columns]
        if missing_cols:
            raise ValueError(f"Missing columns in temporal file: {missing_cols}")
        
        # Create dictionary image_name -> temporal information
        temporal_dict = {}
        for idx, row in temporal_df.iterrows():
            image_name = row['image_name']
            temporal_dict[image_name] = {
                'buildjob': row['buildjob'],
                'geometry_name': row['geometry_name'],
                'channel': row['channel'],
                'layer_index': row['layer_index']
            }
        print(f"Temporal dict created: {len(temporal_dict)} images")
    except Exception as e:
        print(f"ERROR loading temporal information: {e}")
        import traceback
        traceback.print_exc()
        raise
    
    # Load labels (image_name, label, detail)
    print("\nLoading labels...")
    try:
        # Try different engines to avoid issues with Python 3.13
        labels_df = None
        engines_to_try = ['calamine', 'openpyxl', 'xlrd', None]  # calamine is more stable with Python 3.13
        
        for engine in engines_to_try:
            try:
                if engine is None:
                    print(f"   Attempting with default engine...")
                    labels_df = pd.read_excel(labels_path)
                else:
                    print(f"   Attempting with engine '{engine}'...")
                    labels_df = pd.read_excel(labels_path, engine=engine)
                print(f"   Loading completed with engine '{engine if engine else 'default'}'")
                break
            except ImportError:
                print(f"   Engine '{engine}' not available, trying next...")
                continue
            except Exception as e:
                print(f"   Engine '{engine}' failed: {str(e)[:100]}...")
                if engine == engines_to_try[-1]:  # Last attempt
                    raise
                continue
        
        if labels_df is None:
            raise RuntimeError("Unable to load Excel file with any available engine")
        
        print(f"Labels loaded: {len(labels_df)} samples")
        print(f"Columns: {list(labels_df.columns)}")
        
        # Verify required columns
        required_cols = ['image_name', 'label']
        missing_cols = [col for col in required_cols if col not in labels_df.columns]
        if missing_cols:
            raise ValueError(f"❌ Missing columns in labels file: {missing_cols}")
        
        # Verify presence of detail column
        if 'detail' not in labels_df.columns:
            print("WARNING: Column 'detail' not found! Will use 'unknown' for all samples")
            labels_df['detail'] = 'unknown'
        
        print(f"Label distribution: {labels_df['label'].value_counts().to_dict()}")
        print(f"Detail distribution: {labels_df['detail'].value_counts().head(10).to_dict()}")
        
        # Helper function to normalize names (same logic used elsewhere)
        def normalize_name_for_dict(name):
            if pd.isna(name):
                return None
            name_str = str(name).strip()
            for ext in ['.png', '.PNG', '.jpg', '.JPG', '.jpeg', '.JPEG']:
                if name_str.lower().endswith(ext.lower()):
                    name_str = name_str[:-len(ext)]
                    break
            return name_str
        
        # Create dictionaries for fast lookup - USE NORMALIZED NAMES
        labels_dict = {}
        details_dict = {}
        for idx, row in labels_df.iterrows():
            image_name = row['image_name']
            image_name_norm = normalize_name_for_dict(image_name)
            if image_name_norm is not None:
                labels_dict[image_name_norm] = row['label']
                details_dict[image_name_norm] = row['detail'] if pd.notna(row['detail']) else 'unknown'
        
        print(f"Labels dict created: {len(labels_dict)} images (with normalized names)")
    except Exception as e:
        print(f"ERROR loading labels: {e}")
        import traceback
        traceback.print_exc()
        raise
    
    # Load and process ECT data
    print("\nLoading and processing ECT data...")
    print(f"File path: {ect_path}")
    
    try:
        file_size = os.path.getsize(ect_path)
        print(f"File size: {file_size / (1024**2):.2f} MB")
        
        test_file = open(ect_path, 'rb')
        test_file.close()
        print("File accessible and not locked")
    except PermissionError:
        print(f"ERROR: File locked or no permissions: {ect_path}")
        print("   Suggestion: Close Excel if file is open")
        raise
    except Exception as e:
        print(f"Warning during file verification: {e}")
    
    try:
        print("Reading Excel file (this may take a while for large files)...")
        import sys
        import time
        import gc
        sys.stdout.flush()
        
        start_time = time.time()
        
        # Try different engines to avoid issues with Python 3.13
        # calamine is more stable and faster with Python 3.13
        ect_df = None
        
        # Force garbage collection before starting
        gc.collect()
        
        # Attempt 1: calamine (more stable with Python 3.13) - MAXIMUM PRIORITY
        try:
            print("   Attempt 1/5: engine 'calamine' (more stable with Python 3.13)...")
            sys.stdout.flush()
            ect_df = pd.read_excel(ect_path, engine='calamine')
            elapsed = time.time() - start_time
            print(f"   Loading completed in {elapsed:.1f} seconds with engine 'calamine'")
            sys.stdout.flush()
        except ImportError:
            print("   Engine 'calamine' not available, trying next...")
            print("   Suggestion: install with 'pip install python-calamine' for better stability")
            sys.stdout.flush()
        except Exception as e:
            error_msg = str(e)[:200]
            print(f"   Engine 'calamine' failed: {error_msg}...")
            sys.stdout.flush()
            gc.collect()  # Free memory after error
        
        # Attempt 2: openpyxl with read_only=True (more memory efficient)
        if ect_df is None:
            try:
                print("   Attempt 2/5: engine 'openpyxl' with read_only=True (memory efficient)...")
                sys.stdout.flush()
                import openpyxl
                # Use read_only=True to reduce memory usage and crashes
                ect_df = pd.read_excel(ect_path, engine='openpyxl')
                elapsed = time.time() - start_time
                print(f"   Loading completed in {elapsed:.1f} seconds with engine 'openpyxl'")
                sys.stdout.flush()
            except ImportError:
                print("   Engine 'openpyxl' not available, trying next...")
                sys.stdout.flush()
            except Exception as e:
                error_msg = str(e)[:200]
                print(f"   Engine 'openpyxl' failed: {error_msg}...")
                sys.stdout.flush()
                gc.collect()  # Free memory after error
        
        # Attempt 3: openpyxl with load_workbook read_only=True and chunking (workaround for large files)
        if ect_df is None:
            try:
                print("   Attempt 3/5: openpyxl with read_only=True and chunking (for large files)...")
                sys.stdout.flush()
                import openpyxl
                from openpyxl import load_workbook
                import gc
                
                # Load with read_only=True to reduce memory
                print("   Loading workbook in read-only mode...")
                sys.stdout.flush()
                wb = load_workbook(ect_path, read_only=True, data_only=True)
                ws = wb.active
                
                # Read header
                print("   Reading header...")
                sys.stdout.flush()
                first_row = next(ws.iter_rows(values_only=True))
                headers = [str(cell) if cell is not None else f'col_{i}' for i, cell in enumerate(first_row)]
                
                # Read data with chunking to avoid memory accumulation
                print("   Reading data with chunking (this may take time)...")
                sys.stdout.flush()
                chunk_size = 10000  # Process 10k rows at a time
                data_chunks = []
                row_count = 0
                
                for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                    if row_idx == 1:  # Skip already read header
                        continue
                    
                    if any(cell is not None for cell in row):  # Skip completely empty rows
                        data_chunks.append([cell if cell is not None else None for cell in row])
                        row_count += 1
                    
                    # Process chunk periodically to avoid memory accumulation
                    if len(data_chunks) >= chunk_size:
                        if ect_df is None:
                            ect_df = pd.DataFrame(data_chunks, columns=headers)
                        else:
                            chunk_df = pd.DataFrame(data_chunks, columns=headers)
                            ect_df = pd.concat([ect_df, chunk_df], ignore_index=True)
                        data_chunks = []
                        gc.collect()  # Force garbage collection
                        
                        if row_count % 50000 == 0:
                            print(f"   Processed {row_count} rows...")
                            sys.stdout.flush()
                
                # Add last chunk
                if data_chunks:
                    if ect_df is None:
                        ect_df = pd.DataFrame(data_chunks, columns=headers)
                    else:
                        chunk_df = pd.DataFrame(data_chunks, columns=headers)
                        ect_df = pd.concat([ect_df, chunk_df], ignore_index=True)
                
                wb.close()
                del wb, ws, data_chunks
                gc.collect()
                
                elapsed = time.time() - start_time
                print(f"   Loading completed in {elapsed:.1f} seconds with openpyxl chunking ({row_count} rows)")
                sys.stdout.flush()
            except Exception as e:
                error_msg = str(e)[:200]
                print(f"   Openpyxl chunking failed: {error_msg}...")
                sys.stdout.flush()
                gc.collect()  # Free memory after error
        
        # Attempt 4: xlrd
        if ect_df is None:
            try:
                print("   Attempt 4/5: engine 'xlrd'...")
                sys.stdout.flush()
                ect_df = pd.read_excel(ect_path, engine='xlrd')
                elapsed = time.time() - start_time
                print(f"   Loading completed in {elapsed:.1f} seconds with engine 'xlrd'")
                sys.stdout.flush()
            except ImportError:
                print("   Engine 'xlrd' not available, trying next...")
                sys.stdout.flush()
            except Exception as e:
                error_msg = str(e)[:200]
                print(f"   Engine 'xlrd' failed: {error_msg}...")
                sys.stdout.flush()
                gc.collect()  # Free memory after error
        
        # Attempt 5: default engine
        if ect_df is None:
            try:
                print("   Attempt 5/5: default engine...")
                sys.stdout.flush()
                ect_df = pd.read_excel(ect_path)
                elapsed = time.time() - start_time
                print(f"   Loading completed in {elapsed:.1f} seconds with default engine")
                sys.stdout.flush()
            except Exception as e:
                error_msg = str(e)[:200]
                print(f"   Default engine failed: {error_msg}...")
                sys.stdout.flush()
                gc.collect()  # Free memory after error
        
        if ect_df is None:
            raise RuntimeError("Unable to load Excel file with any available engine. "
                             "Try installing 'calamine' with: pip install python-calamine")
        
        # Force final garbage collection
        gc.collect()
        
        print(f"ECT data loaded: {len(ect_df)} datapoints")
        print(f"ECT columns: {list(ect_df.columns)}")
        print(f"Memory usage: {ect_df.memory_usage(deep=True).sum() / (1024**2):.2f} MB")
        sys.stdout.flush()
        
    except MemoryError as e:
        print(f"MEMORY ERROR: File too large to be loaded into memory")
        print(f"   File size: {os.path.getsize(ect_path) / (1024**2):.2f} MB")
        raise
    except Exception as e:
        print(f"ERROR loading ECT data: {type(e).__name__}: {e}")
        import traceback
        traceback.print_exc()
        raise
    
    # Extract features per geometry/channel (original LSTM logic)
    print("\nExtracting features from ECT data...")
    features_df = extract_features_per_geometry(ect_df)
    print(f"Features extracted: {len(features_df)} rows")
    
    # Create centered temporal sequences
    print("\nCreating temporal sequences...")
    ect_sequences, metadata = create_centered_temporal_sequences(features_df, sequence_length=3)
    print(f"Sequences created: {len(ect_sequences)}")
    
    print(f"\nDATA STATISTICS BEFORE MATCHING:")
    print(f"   - Images in labels_df: {len(labels_df)}")
    print(f"   - ECT sequences created: {len(ect_sequences)}")
    
    # Feature selection (as in original LSTM)
    feature_cols = [col for col in features_df.columns if col not in 
                   ['geometry_name', 'channel', 'buildjob', 'layer_index', 'corresponding_layer', 'corresponding_image', 'layer_type', 'Defect']]
    
    # Use center features (all timesteps have the same features)
    # For 3-step sequences: index 0=t-2, 1=t-1, 2=t (center)
    center_features = ect_sequences[:, 2, :]  # Shape: (n_sequences, n_features)
    X = pd.DataFrame(center_features, columns=feature_cols)
    
    # Extract labels for feature selection (from labels file, not from ECT)
    def _normalize_name_fs(name):
        if pd.isna(name):
            return None
        s = str(name).strip()
        for ext in ['.png', '.PNG', '.jpg', '.JPG', '.jpeg', '.JPEG']:
            if s.lower().endswith(ext.lower()):
                s = s[:-len(ext)]
                break
        return s
    labels_dict = { _normalize_name_fs(n): v for n, v in zip(labels_df['image_name'], labels_df['label']) }
    labels_for_fs = []
    for meta in metadata:
        corr_image = meta.get('corresponding_image', None)
        corr_norm = _normalize_name_fs(corr_image)
        if corr_norm in labels_dict:
            label_str = labels_dict[corr_norm]
            labels_for_fs.append(1 if label_str == 'compliant' else 0)
        else:
            labels_for_fs.append(0)  # Default
    
    top_features_idx, top_features_names = feature_selection(X, np.array(labels_for_fs), top_k=16)  # INCREASED from 14 to 16
    
    # Select top features from all sequences
    ect_sequences_filtered = ect_sequences[:, :, top_features_idx]
    ect_input_size = len(top_features_idx)
    
    # Check and clean NaN/Inf in ECT sequences
    nan_count = np.isnan(ect_sequences_filtered).sum()
    inf_count = np.isinf(ect_sequences_filtered).sum()
    if nan_count > 0 or inf_count > 0:
        print(f"WARNING: Found {nan_count} NaN and {inf_count} Inf in ECT sequences")
        print(f"   Replacing with 0...")
        ect_sequences_filtered = np.nan_to_num(ect_sequences_filtered, nan=0.0, posinf=0.0, neginf=0.0)
    
    # Check extreme values
    if np.abs(ect_sequences_filtered).max() > 1e6:
        print(f"WARNING: Extreme values in ECT sequences (max={np.abs(ect_sequences_filtered).max():.2e})")
        print(f"   Applying clipping to [-1000, 1000]...")
        ect_sequences_filtered = np.clip(ect_sequences_filtered, -1000, 1000)
    
    print(f"Selected features: {top_features_names.tolist()}")
    print(f"Final ECT sequence shape: {ect_sequences_filtered.shape}")
    print(f"ECT sequence stats: min={ect_sequences_filtered.min():.4f}, max={ect_sequences_filtered.max():.4f}, mean={ect_sequences_filtered.mean():.4f}, std={ect_sequences_filtered.std():.4f}")
    
    # Create image sequences aligned with ECT sequences
    print("\nCreating image sequences aligned with ECT sequences...")
    image_sequences_metadata, valid_indices = create_image_sequences_aligned_with_ect(
        ect_sequences_filtered, metadata, temporal_dict, labels_dict, details_dict, images_dir
    )
    
    # Filter ECT sequences to keep only aligned ones
    ect_sequences_aligned = ect_sequences_filtered[valid_indices]
    metadata_aligned = [metadata[i] for i in valid_indices]
    
    # NOTE: ECT normalization removed (as in v0/v6) - ECT data is used as is
    
    # Create mapping from original index to filtered index
    # This is necessary because ect_sequence_idx in metadata uses original indices
    original_to_filtered_idx = {original_idx: filtered_idx for filtered_idx, original_idx in enumerate(valid_indices)}
    
    # Update indices in metadata to use filtered indices
    for meta in image_sequences_metadata:
        original_idx = meta['ect_sequence_idx']
        meta['ect_sequence_idx'] = original_to_filtered_idx[original_idx]
    
    print(f"Aligned sequences: {len(image_sequences_metadata)} image sequences, {len(ect_sequences_aligned)} ECT sequences")
    
    # Image transformations - CONSERVATIVE AUGMENTATIONS (from ResNet50_temporal_v2)
    train_transform = transforms.Compose([
        # FORCED resize to fixed size (DOES NOT maintain aspect ratio - faster!)
        transforms.Resize((img_size, img_size)),  # Force fixed size (img_size x img_size)
        
        # CONSERVATIVE augmentations to preserve small defects
        transforms.RandomHorizontalFlip(p=0.5),  # OK - does not hide defects
        transforms.RandomVerticalFlip(p=0.5),    # OK - does not hide defects
        transforms.RandomRotation(degrees=20),   # Kept - OK for LPBF layers
        
        # Color augmentations - kept original for metallic layers
        transforms.ColorJitter(brightness=0.3, contrast=0.3, saturation=0.3, hue=0.15),
        
        # VERY light affine transformation
        transforms.RandomAffine(degrees=0, translate=(0.05, 0.05), scale=(0.95, 1.05)),  # Drastically reduced
        
        # CONSERVATIVE AUGMENTATIONS - REMOVED DANGEROUS ONES
        # REMOVED: RandomPerspective - too much distortion
        # REMOVED: RandomErasing - hides defects!
        
        # Add safe augmentations for texture instead
        transforms.RandomApply([
            transforms.GaussianBlur(kernel_size=3, sigma=(0.1, 0.5))  # Light blur for noise
        ], p=0.2),
        
        transforms.ToTensor(),
        transforms.Normalize([0.485, 0.456, 0.406], [0.229, 0.224, 0.225])
    ])
    
    val_transform = transforms.Compose([
        transforms.Resize((img_size, img_size)),  # Force fixed size (img_size x img_size)
        transforms.ToTensor(),
        transforms.Normalize([0.485, 0.456, 0.406], [0.229, 0.224, 0.225])
    ])
    
    # Create complete dataset with image sequences
    print(f"\nLoading images from: {images_dir}")
    print(f"Checking if directory exists: {os.path.exists(images_dir)}")
    if os.path.exists(images_dir):
        num_files = [f for f in os.listdir(images_dir) if f.lower().endswith(('.png', '.jpg', '.jpeg'))]
        print(f"Found {len(num_files)} image files in directory")
        if len(num_files) > 0:
            print(f"Sample files: {num_files[:3]}")
    else:
        print(f"WARNING: Image directory not found: {images_dir}")
    
    print("\nCreating FusionDataset with image sequences...")
    try:
        full_dataset = FusionDataset(images_dir, image_sequences_metadata, ect_sequences_aligned, transform=None)
        print(f"Dataset created: {len(full_dataset)} samples")
    except Exception as e:
        print(f"ERROR creating dataset: {e}")
        import traceback
        traceback.print_exc()
        raise
    
    # Stratified split - maintains balance between compliant and defective (as ResNet50_temporal_v2)
    print("\nCreating STRATIFIED train/val/test split...")
    from sklearn.model_selection import train_test_split
    
    total_samples = len(full_dataset)
    
    # Extract labels for stratification from sequences
    labels_for_split = []
    for idx in range(total_samples):
        _, _, label, _, _ = full_dataset[idx]
        labels_for_split.append(label)
    
    print(f"Total class distribution: Defective={labels_for_split.count(0)}, Compliant={labels_for_split.count(1)}")
    
    # First split: separate train (80%) from val+test (20%) - STRATIFIED
    train_indices, temp_indices, train_labels, temp_labels = train_test_split(
        list(range(total_samples)), labels_for_split, 
        test_size=0.2, 
        stratify=labels_for_split
        # random_state removed: different split each time, but proportions maintained thanks to stratify
    )
    
    # Second split: separate val (10%) from test (10%) from remaining 20% - STRATIFIED
    # Note: 10% of total = 50% of remaining 20%
    val_indices, test_indices, val_labels, test_labels = train_test_split(
        temp_indices, temp_labels,
        test_size=0.5,  # 50% of 20% = 10% of total
        stratify=temp_labels
        # random_state removed: different split each time, but proportions maintained thanks to stratify
    )
    
    print(f"Stratified split completed:")
    print(f"   Train: {len(train_indices)} samples")
    print(f"   Val: {len(val_indices)} samples")
    print(f"   Test: {len(test_indices)} samples")
    
    # Verify class distribution per split
    train_class_dist = {0: train_labels.count(0), 1: train_labels.count(1)}
    val_class_dist = {0: val_labels.count(0), 1: val_labels.count(1)}
    test_class_dist = {0: test_labels.count(0), 1: test_labels.count(1)}
    print(f"   Train class distribution: defective={train_class_dist[0]}, compliant={train_class_dist[1]}")
    print(f"   Val class distribution: defective={val_class_dist[0]}, compliant={val_class_dist[1]}")
    print(f"   Test class distribution: defective={test_class_dist[0]}, compliant={test_class_dist[1]}")
    
    train_dataset = Subset(full_dataset, train_indices)
    val_dataset = Subset(full_dataset, val_indices)
    test_dataset = Subset(full_dataset, test_indices)
    
    # Apply transformations
    for i in range(len(train_dataset)):
        train_dataset.dataset.transform = train_transform
    for i in range(len(val_dataset)):
        val_dataset.dataset.transform = val_transform
    for i in range(len(test_dataset)):
        test_dataset.dataset.transform = val_transform
    
    # DataLoader - collate_fn for image sequences
    # custom_collate_fn is already defined at module level (before if __name__ == '__main__' block)
    # to allow multiprocessing on Windows
    
    # OPTIMIZATION: DataLoader with persistent_workers and prefetch_factor
    train_loader = DataLoader(train_dataset, batch_size=batch_size, shuffle=True, 
                             collate_fn=custom_collate_fn, num_workers=num_workers, pin_memory=pin_memory,
                             persistent_workers=True if num_workers > 0 else False,
                             prefetch_factor=2 if num_workers > 0 else None)
    val_loader = DataLoader(val_dataset, batch_size=batch_size, shuffle=False, 
                           collate_fn=custom_collate_fn, num_workers=num_workers, pin_memory=pin_memory,
                           persistent_workers=True if num_workers > 0 else False,
                           prefetch_factor=2 if num_workers > 0 else None)
    test_loader = DataLoader(test_dataset, batch_size=batch_size, shuffle=False, 
                            collate_fn=custom_collate_fn, num_workers=num_workers, pin_memory=pin_memory,
                            persistent_workers=True if num_workers > 0 else False,
                            prefetch_factor=2 if num_workers > 0 else None)
    
    print(f"Dataset split: Train: {len(train_dataset)} | Val: {len(val_dataset)} | Test: {len(test_dataset)}")
    
    # Initialize model
    print("\nInitializing model...")
    try:
        model = FusionModel(num_classes=2, dropout_rate=dropout_rate, ect_input_size=ect_input_size, sequence_length=3).to(device)
        
        # OPTIMIZATION: Model compilation to speed up forward pass
        # Check first if Triton is available to avoid errors during training
        model_compiled = False
        if hasattr(torch, 'compile'):
            # Check if Triton is available
            triton_available = False
            triton_version = None
            try:
                import triton
                triton_available = True
                triton_version = getattr(triton, '__version__', 'unknown')
            except ImportError:
                pass
            
            if not triton_available:
                print("Triton not available, skipping compilation (other optimizations still active)")
                print("   To enable compilation on Windows, install: pip install triton-windows")
            else:
                print(f"Triton available (version {triton_version})")
                try:
                    print("Compiling model with torch.compile...")
                    # Save original model state before compiling
                    original_state = model.state_dict()
                    # Try first with 'default', then with other modes if needed
                    model = torch.compile(model, mode='default')
                    # Quick test to verify it works
                    test_input_img = torch.randn(1, 3, 3, 384, 384).to(device)
                    test_input_ect = torch.randn(1, 3, ect_input_size).to(device)
                    with torch.no_grad():
                        _ = model(test_input_img, test_input_ect)
                    print("Model compiled successfully")
                    model_compiled = True
                except Exception as e:
                    error_msg = str(e)
                    print(f"Model compilation failed: {error_msg[:150]}...")
                    print("   Continuing without compilation (other optimizations still active)")
                    # Restore original model if compilation fails
                    model = FusionModel(num_classes=2, dropout_rate=dropout_rate, ect_input_size=ect_input_size, sequence_length=3).to(device)
                    model.load_state_dict(original_state)
        else:
            print("torch.compile not available (requires PyTorch 2.0+)")
        
        print("Model initialized successfully")
    except Exception as e:
        print(f"ERROR initializing model: {e}")
        import traceback
        traceback.print_exc()
        raise
    
    total_params = sum(p.numel() for p in model.parameters())
    trainable_params = sum(p.numel() for p in model.parameters() if p.requires_grad)
    
    print(f"\nModel initialized on: {device}")
    print(f"Total parameters: {total_params:,}")
    print(f"Trainable parameters: {trainable_params:,}")
    
    # Calculate class weights based on DETAIL distribution in dataset
    # OPTIMIZED: uses information already available in sequences instead of loading images
    print("\nCalculating class weights based on DETAIL (optimized)...")
    detail_counts_by_class = {0: {}, 1: {}}  # 0=defective, 1=compliant
    class_counts = {0: 0, 1: 0}
    
    for meta in image_sequences_metadata:
        label = meta['label']
        detail = meta.get('detail', 'unknown')
        class_counts[label] = class_counts.get(label, 0) + 1
        # Also count details for each class
        if detail not in detail_counts_by_class[label]:
            detail_counts_by_class[label][detail] = 0
        detail_counts_by_class[label][detail] += 1
    
    total_samples = len(image_sequences_metadata)
    
    # Calculate weights considering detail distribution
    # Weights are inversely proportional to number of samples per class
    # calculated considering detail distribution
    if 0 in class_counts and 1 in class_counts and class_counts[0] > 0 and class_counts[1] > 0:
        # Weights are calculated based on class counts (which now include details)
        weight_0 = total_samples / (2.0 * class_counts[0])  # defective
        weight_1 = total_samples / (2.0 * class_counts[1])  # compliant
        
        class_weights = torch.tensor([weight_0, weight_1], dtype=torch.float32).to(device)
        print(f"✅ Class weights calculated based on DETAIL in dataset:")
        print(f"  Defective: {weight_0:.3f} (samples: {class_counts[0]}, unique details: {len(detail_counts_by_class[0])})")
        print(f"  Compliant: {weight_1:.3f} (samples: {class_counts[1]}, unique details: {len(detail_counts_by_class[1])})")
        print(f"  Top 5 defective details: {dict(sorted(detail_counts_by_class[0].items(), key=lambda x: x[1], reverse=True)[:5])}")
        print(f"  Top 5 compliant details: {dict(sorted(detail_counts_by_class[1].items(), key=lambda x: x[1], reverse=True)[:5])}")
    else:
        class_weights = None
        print("⚠️  Class weights not calculated (using uniform weights)")
    
    # Loss, Optimizer, Scheduler - WITH DIFFERENTIATED LEARNING RATE
    criterion = nn.CrossEntropyLoss(weight=class_weights, label_smoothing=0.1)  # REDUCED from 0.3 to 0.1
    
    # Differentiated learning rate: TemporalResNet lower (already pretrained), LSTM higher
    image_params = list(model.temporal_resnet.parameters()) + list(model.image_projection.parameters())
    ect_params = list(model.lstm.parameters()) + list(model.attention.parameters()) + list(model.ect_projection.parameters())
    classifier_params = list(model.classifier.parameters())
    
    # Lower LR for TemporalResNet (pretrained), higher for LSTM (new)
    optimizer = optim.AdamW([
        {'params': image_params, 'lr': learning_rate * 0.1},  # TemporalResNet: 10% of base LR
        {'params': ect_params, 'lr': learning_rate * 1.5},    # LSTM: 150% of base LR
        {'params': classifier_params, 'lr': learning_rate}    # Classifier: base LR
    ], weight_decay=weight_decay)
    
    print(f"Differentiated learning rates:")
    print(f"   TemporalResNet (pretrained): {learning_rate * 0.1:.2e}")
    print(f"   LSTM ECT (new): {learning_rate * 1.5:.2e}")
    print(f"   Classifier: {learning_rate:.2e}")
    
    scheduler = optim.lr_scheduler.CosineAnnealingWarmRestarts(optimizer, T_0=8, T_mult=2)
    early_stopping = EarlyStopping(patience=5, min_delta=0.005)  # NOT MODIFIED
    
    # OPTIMIZATION: Mixed precision training (FP16)
    scaler = torch.cuda.amp.GradScaler() if device_type == "gpu" else None
    use_amp = scaler is not None
    if use_amp:
        print("🚀 Mixed precision training (FP16) enabled")
    else:
        print("⚠️  Mixed precision not available (CPU mode or CUDA does not support FP16)")
    
    # Training loop
    print(f"\n{'='*80}")
    print(f"STARTING TRAINING - {device_type.upper()} MODE")
    print(f"{'='*80}")
    
    best_val_acc = 0
    train_losses = []
    val_losses = []
    train_accs = []
    val_accs = []
    
    # Gradient clipping
    max_grad_norm = 1.0
    
    for epoch in range(num_epochs):
        # Training
        model.train()
        running_loss = 0.0
        correct = 0
        total = 0
        
        train_pbar = tqdm(train_loader, desc=f"Epoch {epoch+1}/{num_epochs} [Train]", 
                          leave=False, ncols=120, colour='green')
        
        for batch_idx, (image_sequences, ect_seqs, labels, center_images, details) in enumerate(train_pbar):
            image_sequences, ect_seqs, labels = image_sequences.to(device, non_blocking=True), ect_seqs.to(device, non_blocking=True), labels.to(device, non_blocking=True)
            
            # OPTIMIZATION: NaN/Inf checks removed from training (only in validation)
            # Data is already cleaned in dataset, so no checks needed here
            
            optimizer.zero_grad()
            
            # OPTIMIZATION: Mixed precision forward pass
            if use_amp:
                with torch.cuda.amp.autocast():
                    outputs = model(image_sequences, ect_seqs)
                    loss = criterion(outputs, labels)
            else:
                outputs = model(image_sequences, ect_seqs)
                loss = criterion(outputs, labels)
            
            # OPTIMIZATION: Mixed precision backward pass
            if use_amp:
                scaler.scale(loss).backward()
                # Gradient clipping with scaler
                scaler.unscale_(optimizer)
                torch.nn.utils.clip_grad_norm_(model.parameters(), max_grad_norm)
                scaler.step(optimizer)
                scaler.update()
            else:
                loss.backward()
                # Gradient clipping
                torch.nn.utils.clip_grad_norm_(model.parameters(), max_grad_norm)
                optimizer.step()
            
            running_loss += loss.item() * image_sequences.size(0)
            _, predicted = torch.max(outputs, 1)
            total += labels.size(0)
            correct += (predicted == labels).sum().item()
            
            train_pbar.set_postfix({
                'Loss': f'{loss.item():.4f}',
                'Acc': f'{100.*correct/total:.1f}%',
                'LR': f'{optimizer.param_groups[0]["lr"]:.2e}'
            })
        
        train_loss = running_loss / total
        train_acc = correct / total
        train_losses.append(train_loss)
        train_accs.append(train_acc)
        
        # Validation
        model.eval()
        val_loss = 0.0
        val_correct = 0
        val_total = 0
        
        val_pbar = tqdm(val_loader, desc=f"Epoch {epoch+1}/{num_epochs} [Val]", 
                        leave=False, ncols=120, colour='blue')
        
        with torch.no_grad():
            for image_sequences, ect_seqs, labels, center_images, details in val_pbar:
                image_sequences, ect_seqs, labels = image_sequences.to(device, non_blocking=True), ect_seqs.to(device, non_blocking=True), labels.to(device, non_blocking=True)
                
                # OPTIMIZATION: NaN/Inf checks only in validation (for debug)
                if torch.isnan(image_sequences).any() or torch.isnan(ect_seqs).any():
                    continue
                if torch.isinf(image_sequences).any() or torch.isinf(ect_seqs).any():
                    continue
                
                # OPTIMIZATION: Mixed precision also in validation
                if use_amp:
                    with torch.cuda.amp.autocast():
                        outputs = model(image_sequences, ect_seqs)
                        loss = criterion(outputs, labels)
                else:
                    outputs = model(image_sequences, ect_seqs)
                    loss = criterion(outputs, labels)
                
                # Check NaN in output (only in validation)
                if torch.isnan(outputs).any():
                    continue
                
                # Check NaN in loss (only in validation)
                if torch.isnan(loss) or torch.isinf(loss):
                    continue
                
                val_loss += loss.item() * image_sequences.size(0)
                _, predicted = torch.max(outputs, 1)
                val_total += labels.size(0)
                val_correct += (predicted == labels).sum().item()
                
                val_pbar.set_postfix({
                    'Loss': f'{loss.item():.4f}',
                    'Acc': f'{100.*val_correct/val_total:.1f}%'
                })
        
        val_loss /= val_total
        val_acc = val_correct / val_total
        val_losses.append(val_loss)
        val_accs.append(val_acc)
        
        scheduler.step()
        
        gap = train_acc - val_acc
        
        if early_stopping(val_loss, val_acc, model):
            print(f"\nEarly stopping triggered at epoch {epoch+1}")
            break
        
        if gap > 0.03:
            print(f"Gap too large ({gap:.4f}), considering early stop...")
        
        if val_acc > best_val_acc:
            best_val_acc = val_acc
            # Make sure directory exists before saving
            os.makedirs(results_dir, exist_ok=True)
            model_filename = os.path.join(results_dir, f'best_fusion_model_v7_temporal_3steps_balanced_{device_type}.pth')
            torch.save({
                'epoch': epoch,
                'model_state_dict': model.state_dict(),
                'optimizer_state_dict': optimizer.state_dict(),
                'val_acc': val_acc,
                'val_loss': val_loss,
                'device_type': device_type,
                'params': params
            }, model_filename)
        
        print(f"Epoch [{epoch+1:2d}/{num_epochs}] "
              f"Train Loss: {train_loss:.4f} Train Acc: {train_acc:.4f} "
              f"Val Loss: {val_loss:.4f} Val Acc: {val_acc:.4f} "
              f"Gap: {gap:.4f} LR: {optimizer.param_groups[0]['lr']:.2e}")
    
    print(f"\nTraining completed! Best validation accuracy: {best_val_acc:.4f}")
    
    # Save training summary
    print("\nSaving training summary...")
    training_summary_filename = os.path.join(results_dir, f"training_summary_v7_temporal_3steps_balanced_{device_type}.txt")
    with open(training_summary_filename, "w", encoding="utf-8") as f:
        f.write("="*80 + "\n")
        f.write("TRAINING SUMMARY - FUSION MODEL V7 TEMPORAL BALANCED (3 STEPS)\n")
        f.write("="*80 + "\n\n")
        f.write(f"Device: {device_type.upper()}\n")
        f.write(f"Total epochs: {len(train_losses)}\n")
        f.write(f"Best validation accuracy: {best_val_acc:.4f}\n")
        f.write("="*80 + "\n\n")
        
        f.write(f"{'Epoch':<8} {'Train Loss':<12} {'Train Acc':<12} {'Val Loss':<12} {'Val Acc':<12}\n")
        f.write("-"*80 + "\n")
        
        for epoch in range(len(train_losses)):
            f.write(f"{epoch+1:<8} {train_losses[epoch]:<12.6f} {train_accs[epoch]:<12.6f} "
                   f"{val_losses[epoch]:<12.6f} {val_accs[epoch]:<12.6f}\n")
    
    # Make sure directory exists before saving
    os.makedirs(results_dir, exist_ok=True)
    training_summary_csv = os.path.join(results_dir, f"training_summary_v7_temporal_3steps_balanced_{device_type}.csv")
    df_summary = pd.DataFrame({
        'epoch': range(1, len(train_losses) + 1),
        'train_loss': train_losses,
        'train_acc': train_accs,
        'val_loss': val_losses,
        'val_acc': val_accs
    })
    df_summary.to_csv(training_summary_csv, index=False)
    print(f"Training summary saved: {training_summary_filename}")
    
    # Final test
    print("\nStarting final test...")
    model_filename = os.path.join(results_dir, f'best_fusion_model_v7_temporal_3steps_balanced_{device_type}.pth')
    checkpoint = torch.load(model_filename)
    model.load_state_dict(checkpoint['model_state_dict'])
    print(f"Best model loaded: {model_filename}")
    
    model.eval()
    test_correct = 0
    test_total = 0
    y_true = []
    y_pred = []
    y_probs = []
    y_details = []
    misclassified_images = []
    
    test_pbar = tqdm(test_loader, desc="Final Test", ncols=120, colour='red')
    
    with torch.no_grad():
        for image_sequences, ect_seqs, labels, center_images, details in test_pbar:
            image_sequences, ect_seqs, labels = image_sequences.to(device, non_blocking=True), ect_seqs.to(device, non_blocking=True), labels.to(device, non_blocking=True)
            
            # OPTIMIZATION: Mixed precision also in test
            if use_amp:
                with torch.cuda.amp.autocast():
                    outputs = model(image_sequences, ect_seqs)
            else:
                outputs = model(image_sequences, ect_seqs)
            probs = torch.softmax(outputs, dim=1)
            _, predicted = torch.max(outputs, 1)
            
            test_total += labels.size(0)
            test_correct += (predicted == labels).sum().item()
            
            for i in range(len(labels)):
                true_label = labels[i].item()
                pred_label = predicted[i].item()
                center_image = center_images[i]
                # Convert detail to Python value (could be tensor or already value)
                detail_raw = details[i]
                if isinstance(detail_raw, torch.Tensor):
                    if detail_raw.numel() == 1:
                        detail = detail_raw.item()
                        # Handle NaN from tensor
                        if isinstance(detail, float) and np.isnan(detail):
                            detail = 'N/A'
                        else:
                            detail = str(detail) if not isinstance(detail, (str, int, float)) else detail
                    else:
                        detail = str(detail_raw.item())
                elif detail_raw is None:
                    detail = 'N/A'
                elif isinstance(detail_raw, float) and np.isnan(detail_raw):
                    detail = 'N/A'
                elif pd.isna(detail_raw):
                    detail = 'N/A'
                else:
                    detail = str(detail_raw) if not isinstance(detail_raw, (str, int, float)) else detail_raw
                prob = probs[i, 1].item()
                
                y_true.append(true_label)
                y_pred.append(pred_label)
                y_probs.append(prob)
                y_details.append(detail)
                
                if true_label != pred_label:
                    true_class = 'compliant' if true_label == 1 else 'defective'
                    pred_class = 'compliant' if pred_label == 1 else 'defective'
                    # Find full path of central image
                    center_image_path = None
                    center_image_base = str(center_image).split('.')[0]
                    search_paths = [
                        os.path.join(IMAGES_DIR, center_image),
                        os.path.join(IMAGES_DIR, center_image + '.png'),
                        os.path.join(IMAGES_DIR, center_image + '.PNG'),
                        os.path.join(IMAGES_DIR, center_image_base + '.png'),
                        os.path.join(IMAGES_DIR, center_image_base + '.PNG'),
                    ]
                    for search_path in search_paths:
                        if os.path.exists(search_path):
                            center_image_path = search_path
                            break
                    if center_image_path is None:
                        # Search by prefix
                        for file in os.listdir(IMAGES_DIR):
                            if file.lower().startswith(center_image_base.lower()):
                                center_image_path = os.path.join(IMAGES_DIR, file)
                                break
                    if center_image_path is None:
                        center_image_path = os.path.join(IMAGES_DIR, center_image)
                    
                    misclassified_images.append({
                        'center_image_name': center_image,
                        'true_class': true_class,
                        'predicted_class': pred_class,
                        'detail': detail,
                        'confidence': prob if pred_label == 1 else 1-prob,
                        'path': center_image_path
                    })
            
            test_pbar.set_postfix({
                'Acc': f'{100.*test_correct/test_total:.1f}%',
                'Errors': len(misclassified_images)
            })
    
    test_accuracy = test_correct / test_total
    print(f"\nTest Accuracy: {test_accuracy:.4f}")
    
    # Advanced metrics
    precision, recall, f1, _ = precision_recall_fscore_support(y_true, y_pred, average='weighted')
    auc = roc_auc_score(y_true, y_probs)
    
    print(f"\n{'='*80}")
    print(f"ADVANCED METRICS - {device_type.upper()} MODE")
    print(f"{'='*80}")
    print(f"Accuracy:  {test_accuracy:.4f}")
    print(f"Precision: {precision:.4f}")
    print(f"Recall:    {recall:.4f}")
    print(f"F1-Score:  {f1:.4f}")
    print(f"AUC:       {auc:.4f}")
    
    # Add final metrics to training summary
    with open(training_summary_filename, "a", encoding="utf-8") as f:
        f.write("\n" + "="*80 + "\n")
        f.write("FINAL PERFORMANCE METRICS - TEST SET\n")
        f.write("="*80 + "\n\n")
        f.write(f"Test Accuracy:  {test_accuracy:.6f} ({test_accuracy*100:.2f}%)\n")
        f.write(f"Precision:      {precision:.6f} ({precision*100:.2f}%)\n")
        f.write(f"Recall:         {recall:.6f} ({recall*100:.2f}%)\n")
        f.write(f"F1-Score:       {f1:.6f} ({f1*100:.2f}%)\n")
        f.write(f"AUC (ROC):      {auc:.6f}\n")
        f.write(f"\nTotal test images: {test_total}\n")
        f.write(f"Correctly classified images: {test_correct}\n")
        f.write(f"Misclassified images: {len(misclassified_images)}\n")
        f.write("="*80 + "\n")
    
    # Make sure directory exists before saving
    os.makedirs(results_dir, exist_ok=True)
    final_metrics_csv = os.path.join(results_dir, f"final_metrics_v7_temporal_3steps_balanced_{device_type}.csv")
    df_final = pd.DataFrame({
        'metric': ['test_accuracy', 'test_precision', 'test_recall', 'test_f1', 'test_auc', 'test_total', 'test_correct', 'test_errors'],
        'value': [test_accuracy, precision, recall, f1, auc, test_total, test_correct, len(misclassified_images)]
    })
    df_final.to_csv(final_metrics_csv, index=False)
    
    # -------------------------------
    # PRINT MISCLASSIFIED IMAGES
    # -------------------------------
    print(f"\n{'='*80}")
    print(f"MISCLASSIFIED SEQUENCES ({len(misclassified_images)} out of {test_total})")
    print(f"{'='*80}")
    
    if misclassified_images:
        # Sort by confidence (more confident = more serious errors)
        misclassified_images.sort(key=lambda x: x['confidence'], reverse=True)
        
        for i, item in enumerate(misclassified_images, 1):
            print(f"{i:3d}. {item['center_image_name']}")
            print(f"   True: {item['true_class']} | Predicted: {item['predicted_class']}")
            print(f"   Detail: {item.get('detail', 'N/A')}")
            print(f"   Confidence: {item['confidence']:.3f}")
            print(f"   Path: {item.get('path', 'N/A')}")
            print()
    else:
        print("No misclassified sequences!")
    
    # -------------------------------
    # SAVE FILES AND MISCLASSIFIED IMAGES
    # -------------------------------
    # Create folder for misclassified images
    misclassified_dir = os.path.join(results_dir, f"misclassified_sequences_v7_temporal_3steps_balanced_{device_type}")
    if not os.path.exists(misclassified_dir):
        os.makedirs(misclassified_dir)
        print(f"Folder created: {misclassified_dir}")
    
    # Make sure directory exists before saving
    os.makedirs(results_dir, exist_ok=True)
    # Save error list in Excel format
    error_filename = os.path.join(results_dir, f"misclassified_sequences_v7_temporal_3steps_balanced_{device_type}.xlsx")
    print(f"Saving error list in: {error_filename}")
    
    if misclassified_images:
        # Create DataFrame with required columns
        error_data = []
        for i, item in enumerate(misclassified_images, 1):
            error_data.append({
                'image_name': item['center_image_name'],
                'true': item['true_class'],
                'predicted': item['predicted_class'],
                'detail': item.get('detail', 'N/A'),
                'confidence': item['confidence'],
                'path': item.get('path', 'N/A')
            })
            
            # Copy central image to misclassified folder
            try:
                center_image_name = item['center_image_name']
                center_image_path = item.get('path', os.path.join(IMAGES_DIR, center_image_name))
                
                # Create filename with additional information
                base_name = os.path.splitext(center_image_name)[0]
                extension = os.path.splitext(center_image_name)[1]
                if not extension:
                    extension = '.png'  # Default extension
                new_filename = f"{i:03d}_{base_name}_TRUE_{item['true_class']}_PRED_{item['predicted_class']}_CONF_{item['confidence']:.3f}{extension}"
                
                # Copy image if path exists
                if os.path.exists(center_image_path):
                    shutil.copy2(center_image_path, os.path.join(misclassified_dir, new_filename))
                else:
                    print(f"Warning: Image not found: {center_image_path}")
            except Exception as e:
                print(f"Error copying {item['center_image_name']}: {str(e)}")
        
        # Create DataFrame and save as Excel
        df_errors = pd.DataFrame(error_data)
        df_errors.to_excel(error_filename, index=False, engine='openpyxl')
        print(f"Excel file created with {len(df_errors)} rows")
    else:
        # Create empty DataFrame with correct columns
        df_errors = pd.DataFrame(columns=['image_name', 'true', 'predicted', 'detail', 'confidence', 'path'])
        df_errors.to_excel(error_filename, index=False, engine='openpyxl')
        print("No misclassified sequences! Empty Excel file created.")
    
    print(f"Misclassified images copied to: {misclassified_dir}")
    
    # Confusion Matrix
    print("\nGenerating confusion matrix...")
    cm = confusion_matrix(y_true, y_pred)
    class_names = ['defective', 'compliant']
    
    plt.figure(figsize=(10, 8))
    sns.heatmap(cm, annot=True, fmt="d", cmap="Blues",
                xticklabels=class_names, yticklabels=class_names)
    plt.xlabel("Predicted")
    plt.ylabel("True")
    plt.title(f"Confusion Matrix - Fusion Model V7 Temporal Balanced (3 Steps) - {device_type.upper()} - Accuracy: {test_accuracy:.4f}")
    plt.tight_layout()
    cm_filename = os.path.join(results_dir, f"confusion_matrix_v7_temporal_3steps_balanced_{device_type}.png")
    plt.savefig(cm_filename, dpi=300, bbox_inches='tight')
    plt.close()
    
    print("\nClassification Report:")
    print(classification_report(y_true, y_pred, target_names=class_names))
    
    # ANALYSIS BY DETAIL TYPE
    print("\n" + "="*80)
    print("ANALYSIS BY DETAIL TYPE")
    print("="*80)
    
    detail_groups = {}
    for i, detail in enumerate(y_details):
        # Normalize detail: convert to string and handle None/NaN values
        if detail is None or (isinstance(detail, float) and np.isnan(detail)):
            detail_normalized = 'N/A'
        else:
            detail_normalized = str(detail).strip()
        
        if detail_normalized not in detail_groups:
            detail_groups[detail_normalized] = {'y_true': [], 'y_pred': [], 'y_probs': []}
        detail_groups[detail_normalized]['y_true'].append(y_true[i])
        detail_groups[detail_normalized]['y_pred'].append(y_pred[i])
        detail_groups[detail_normalized]['y_probs'].append(y_probs[i])
    
    sorted_details = sorted(detail_groups.items(), key=lambda x: len(x[1]['y_true']), reverse=True)
    
    print(f"\nFound {len(detail_groups)} detail types in test set")
    print(f"Sample distribution per detail:")
    for detail, data in sorted_details:
        print(f"  - {detail}: {len(data['y_true'])} samples")
    
    detail_metrics_list = []
    confusion_matrices_created = 0
    
    for detail, data in sorted_details:
        if len(data['y_true']) == 0:
            print(f"⚠️  Skipped detail '{detail}': no samples")
            continue
        
        detail_y_true = data['y_true']
        detail_y_pred = data['y_pred']
        detail_y_probs = data['y_probs']
        
        detail_accuracy = sum(1 for i in range(len(detail_y_true)) if detail_y_true[i] == detail_y_pred[i]) / len(detail_y_true)
        
        try:
            detail_precision, detail_recall, detail_f1, _ = precision_recall_fscore_support(
                detail_y_true, detail_y_pred, average='weighted', zero_division=0
            )
        except:
            detail_precision = detail_recall = detail_f1 = 0.0
        
        try:
            if len(set(detail_y_true)) > 1:
                detail_auc = roc_auc_score(detail_y_true, detail_y_probs)
            else:
                detail_auc = float('nan')
        except:
            detail_auc = float('nan')
        
        detail_metrics_list.append({
            'detail': detail,
            'n_samples': len(detail_y_true),
            'accuracy': detail_accuracy,
            'precision': detail_precision,
            'recall': detail_recall,
            'f1_score': detail_f1,
            'auc': detail_auc
        })
        
        # Confusion matrix per detail
        detail_cm = confusion_matrix(detail_y_true, detail_y_pred)
        
        plt.figure(figsize=(10, 8))
        sns.heatmap(detail_cm, annot=True, fmt="d", cmap="Blues",
                    xticklabels=class_names, yticklabels=class_names)
        plt.xlabel("Predicted")
        plt.ylabel("True")
        plt.title(f"Confusion Matrix - Detail: {detail}\nAccuracy: {detail_accuracy:.4f} | Samples: {len(detail_y_true)}")
        plt.tight_layout()
        
        safe_detail_name = str(detail)
        invalid_chars = ['<', '>', ':', '"', '/', '\\', '|', '?', '*']
        for char in invalid_chars:
            safe_detail_name = safe_detail_name.replace(char, '_')
        safe_detail_name = '_'.join(safe_detail_name.split())
        if len(safe_detail_name) > 100:
            safe_detail_name = safe_detail_name[:100]
        detail_cm_filename = os.path.join(results_dir, f"confusion_matrix_detail_{safe_detail_name}_{device_type}.png")
        try:
            plt.savefig(detail_cm_filename, dpi=300, bbox_inches='tight')
            plt.close()
            confusion_matrices_created += 1
            print(f"✅ Confusion matrix created for detail '{detail}': {detail_cm_filename}")
        except Exception as e:
            print(f"❌ Error saving confusion matrix for detail '{detail}': {e}")
            plt.close()
        
        print(f"\nDetail: {detail}")
        print(f"  Samples: {len(detail_y_true)}")
        print(f"  Accuracy: {detail_accuracy:.4f}")
        print(f"  Precision: {detail_precision:.4f}")
        print(f"  Recall: {detail_recall:.4f}")
        print(f"  F1-Score: {detail_f1:.4f}")
        auc_str = f"{detail_auc:.4f}" if not np.isnan(detail_auc) else "N/A"
        print(f"  AUC: {auc_str}")
    
    # Make sure directory exists before saving
    os.makedirs(results_dir, exist_ok=True)
    df_detail_metrics = pd.DataFrame(detail_metrics_list)
    detail_metrics_csv = os.path.join(results_dir, f"detail_metrics_v7_temporal_3steps_balanced_{device_type}.csv")
    df_detail_metrics.to_csv(detail_metrics_csv, index=False)
    print(f"\n{'='*80}")
    print(f"SUMMARY OF CONFUSION MATRICES PER DETAIL")
    print(f"{'='*80}")
    print(f"Total details found: {len(detail_groups)}")
    print(f"Details with samples: {len([d for d, data in sorted_details if len(data['y_true']) > 0])}")
    print(f"Confusion matrices created: {confusion_matrices_created}")
    print(f"{'='*80}")
    print(f"\nDetail metrics saved in: {detail_metrics_csv}")
    
    # Add detail metrics to training summary
    with open(training_summary_filename, "a", encoding="utf-8") as f:
        f.write("\n" + "="*80 + "\n")
        f.write("METRICS BY DETAIL TYPE - TEST SET\n")
        f.write("="*80 + "\n\n")
        
        for detail, data in sorted_details:
            if len(data['y_true']) == 0:
                continue
            
            detail_y_true = data['y_true']
            detail_y_pred = data['y_pred']
            detail_y_probs = data['y_probs']
            
            detail_accuracy = sum(1 for i in range(len(detail_y_true)) if detail_y_true[i] == detail_y_pred[i]) / len(detail_y_true)
            
            try:
                detail_precision, detail_recall, detail_f1, _ = precision_recall_fscore_support(
                    detail_y_true, detail_y_pred, average='weighted', zero_division=0
                )
            except:
                detail_precision = detail_recall = detail_f1 = 0.0
            
            try:
                if len(set(detail_y_true)) > 1:
                    detail_auc = roc_auc_score(detail_y_true, detail_y_probs)
                else:
                    detail_auc = float('nan')
            except:
                detail_auc = float('nan')
            
            f.write(f"Detail: {detail}\n")
            f.write(f"  Samples: {len(detail_y_true)}\n")
            f.write(f"  Accuracy:  {detail_accuracy:.6f} ({detail_accuracy*100:.2f}%)\n")
            f.write(f"  Precision: {detail_precision:.6f} ({detail_precision*100:.2f}%)\n")
            f.write(f"  Recall:    {detail_recall:.6f} ({detail_recall*100:.2f}%)\n")
            f.write(f"  F1-Score:  {detail_f1:.6f} ({detail_f1*100:.2f}%)\n")
            if not np.isnan(detail_auc):
                f.write(f"  AUC (ROC): {detail_auc:.6f}\n")
            else:
                f.write(f"  AUC (ROC): N/A\n")
            f.write("\n")
        
        f.write("="*80 + "\n")
    
    # Training Progress Plot
    plt.figure(figsize=(20, 6))
    
    plt.subplot(1, 3, 1)
    plt.plot(train_losses, label='Train Loss', color='blue')
    plt.plot(val_losses, label='Val Loss', color='red')
    plt.xlabel('Epoch')
    plt.ylabel('Loss')
    plt.title(f'Training Progress - Loss V7 ({device_type.upper()})')
    plt.legend()
    plt.grid(True)
    
    plt.subplot(1, 3, 2)
    plt.plot(train_accs, label='Train Acc', color='blue')
    plt.plot(val_accs, label='Val Acc', color='red')
    plt.xlabel('Epoch')
    plt.ylabel('Accuracy')
    plt.title(f'Training Progress - Accuracy V7 ({device_type.upper()})')
    plt.legend()
    plt.grid(True)
    
    plt.subplot(1, 3, 3)
    epochs_range = range(1, len(train_losses) + 1)
    lrs = [optimizer.param_groups[0]['lr'] for _ in epochs_range]
    plt.plot(epochs_range, lrs, label='Learning Rate', color='green')
    plt.xlabel('Epoch')
    plt.ylabel('Learning Rate')
    plt.title('Learning Rate Schedule')
    plt.legend()
    plt.grid(True)
    
    plt.tight_layout()
    progress_filename = os.path.join(results_dir, f"training_progress_v7_balanced_{device_type}.png")
    plt.savefig(progress_filename, dpi=300, bbox_inches='tight')
    plt.close()
    
    print("\nAll completed!")
    print(f"All outputs saved in: {results_dir}")
    print(f"\nSaved files:")
    print(f"  - Model: {model_filename}")
    print(f"  - Training summary (txt): {training_summary_filename}")
    print(f"  - Training summary (csv): {training_summary_csv}")
    print(f"  - Final metrics (csv): {final_metrics_csv}")
    print(f"  - Detail metrics (csv): {detail_metrics_csv}")
    print(f"  - Plots: {cm_filename}, {progress_filename}")
    print(f"  - Confusion matrix per detail: {len(detail_groups)} files")
    print(f"  - Errors (Excel): {error_filename}")
    print(f"  - Misclassified images: {misclassified_dir}")
    print(f"\n{'='*80}")



if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n\n⚠️  Manual interruption by user")
    except Exception as e:
        print(f"\n\n❌ CRITICAL ERROR: {type(e).__name__}: {str(e)}")
        import traceback
        print("\n📋 Full traceback:")
        traceback.print_exc()
        input("\nPress ENTER to close...")

